import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from copy import copy
import re
import unicodedata

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================
# Configuración básica
# =============================
st.set_page_config(page_title="Procesador de Exp Contable - SIAF", layout="wide")
st.title("Procesador de Exp Contable - SIAF")

# Constantes
REQUIRED_FOR_EXP = ["ano_eje", "nro_not_exp", "ciclo", "fase"]
NUMERIC_COLS = ["debe", "haber", "saldo"]
EQUIV_SHEET = "Hoja de Trabajo"
COPIABLE_SHEET = "HT EF-4"
MISSING_RUBRO_LABEL = "(Sin Rubro)"
MISSING_EF4_LABEL = "(Sin EF-4)"
MISSING_ACT_LABEL = "(Sin Actividad)"

# =============================
# Utilidades / Helpers
# =============================
@st.cache_data(show_spinner=False)
def _read_file_bytes(uploaded_file) -> bytes:
    return uploaded_file.read()

@st.cache_data(show_spinner=False)
def load_main_df(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(BytesIO(file_bytes), dtype=str, engine="openpyxl")
    df.columns = [c.strip().lower() for c in df.columns]

    # Coerción numérica segura
    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    # Validación de columnas requeridas
    missing = [c for c in REQUIRED_FOR_EXP if c not in df.columns]
    if missing:
        raise ValueError(f"Faltan columnas requeridas en el archivo principal: {', '.join(missing)}")

    # Construcción de exp_contable
    parts = [df[c].astype(str).fillna("") for c in REQUIRED_FOR_EXP]
    df["exp_contable"] = parts[0] + "-" + parts[1] + "-" + parts[2] + "-" + parts[3]

    # Clave contable
    mayor = df["mayor"].astype(str) if "mayor" in df.columns else pd.Series("", index=df.index)
    sub_cta = df["sub_cta"].astype(str) if "sub_cta" in df.columns else pd.Series("", index=df.index)
    df["clave_cta"] = mayor.str.strip() + "." + sub_cta.str.strip()

    return df

@st.cache_data(show_spinner=False)
def load_equiv_df(file_bytes: bytes) -> pd.DataFrame:
    # Siempre “Hoja de Trabajo” del archivo de equivalencias
    try:
        df_e = pd.read_excel(BytesIO(file_bytes), sheet_name=EQUIV_SHEET, dtype=str, engine="openpyxl")
    except ValueError as e:
        raise ValueError(f"No se encontró la hoja '{EQUIV_SHEET}' en el archivo de Equivalencias.") from e

    df_e.columns = [str(c).strip() for c in df_e.columns]
    required_cols = {"Cuentas Contables", "Rubros"}
    if not required_cols.issubset(df_e.columns):
        raise ValueError("La hoja de Equivalencias debe contener las columnas 'Cuentas Contables' y 'Rubros'.")

    # No eliminamos columnas extra (usamos EF-4 y Actividad si existen)
    df_e["Cuentas Contables"] = df_e["Cuentas Contables"].astype(str).str.strip()
    df_e["Rubros"] = df_e["Rubros"].astype(str).str.strip()
    df_e = df_e.drop_duplicates(subset=["Cuentas Contables"], keep="first").reset_index(drop=True)
    return df_e

def compute_adjusted(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ajuste debe/haber:
      - Swap (debe<->haber) SOLO para expedientes que NO son mayor 1101 (-> irá a Tipo1_sin_1101).
      - Para mayor == 1101 (Tipo1_con_1101) se mantiene debe/haber tal cual.
    """
    mayor_series = df["mayor"].astype(str) if "mayor" in df.columns else pd.Series("", index=df.index)
    exp_con_1101 = set(df.loc[mayor_series.eq("1101"), "exp_contable"].unique())
    in_1101 = df["exp_contable"].isin(exp_con_1101)  # True si el expediente es 1101

    debe = pd.to_numeric(df.get("debe", 0), errors="coerce").fillna(0.0)
    haber = pd.to_numeric(df.get("haber", 0), errors="coerce").fillna(0.0)

    # Si es 1101: mantener; si no: invertir
    df["debe_adj"]  = np.where(in_1101, debe,  haber)
    df["haber_adj"] = np.where(in_1101, haber, debe)

    # Tipos finales
    for col in ["debe_adj", "haber_adj", "debe", "haber", "saldo"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    for col in df.columns:
        if col not in ["debe_adj", "haber_adj", "debe", "haber", "saldo"]:
            df[col] = df[col].astype(str)

    return df

def merge_equivalences(df: pd.DataFrame, df_equiv: pd.DataFrame) -> pd.DataFrame:
    return df.merge(
        df_equiv[["Cuentas Contables", "Rubros"]],
        left_on="clave_cta",
        right_on="Cuentas Contables",
        how="left",
    )

def copy_sheet_with_styles(src_ws: openpyxl.worksheet.worksheet.Worksheet,
                           dst_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

def is_inside_merged_area(row: int, col: int, merged_ranges) -> bool:
    for rng in merged_ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return True
    return False

# ---------- Normalizadores / búsqueda tolerante ----------
def _norm_text(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s)).encode("ascii", "ignore").decode("ascii")
    s = s.replace("_", " ").replace("-", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

def _norm_account_code(s: str) -> str:
    s = str(s or "")
    s = s.strip()
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^0-9.]", "", s)
    s = re.sub(r"\.+", ".", s)
    return s.strip(".")

def _pick_col(df: pd.DataFrame, candidates: list[str], must_contain: str | None = None) -> str | None:
    cand_norm = {_norm_text(c) for c in candidates}
    for c in df.columns:
        if _norm_text(c) in cand_norm:
            return c
    if must_contain:
        mc = _norm_text(must_contain)
        for c in df.columns:
            if mc in _norm_text(c):
                return c
    return None

def _find_sheet_name(equiv_bytes: bytes, patterns: list[str]) -> str | None:
    wb = openpyxl.load_workbook(BytesIO(equiv_bytes), read_only=True, data_only=True)
    norm_map = {name: _norm_text(name) for name in wb.sheetnames}
    for name, n in norm_map.items():
        for pat in patterns:
            if re.search(pat, n):
                return name
    return None

def _find_structure_sheet_name(equiv_bytes: bytes) -> str | None:
    """Encuentra la hoja de estructura de Rubros de forma robusta."""
    wb = openpyxl.load_workbook(BytesIO(equiv_bytes), read_only=True, data_only=True)
    exact_norm = {"estructura del archivo", "estructura_del_archivo"}
    candidates = []
    for name in wb.sheetnames:
        n = _norm_text(name)
        if n in exact_norm:
            return name
        if "estruct" in n:
            candidates.append(name)
    # Validar candidatos por columnas compatibles
    for name in candidates:
        try:
            df_test = pd.read_excel(BytesIO(equiv_bytes), sheet_name=name, dtype=str, engine="openpyxl", nrows=5)
            cols_norm = [_norm_text(c) for c in df_test.columns]
            if any(
                cn in {"estructura", "descripcion", "descripción", "rubros", "rubro"} or
                ("estruct" in cn) or ("descr" in cn)
                for cn in cols_norm
            ):
                return name
        except Exception:
            pass
    return None

def _find_estructura_ef4_sheet_name(equiv_bytes: bytes) -> str | None:
    """Encuentra la hoja 'Estructura EF-4' (o similar) para ordenar EF-4 y Actividades."""
    wb = openpyxl.load_workbook(BytesIO(equiv_bytes), read_only=True, data_only=True)
    for name in wb.sheetnames:
        n = _norm_text(name)
        if "estruct" in n and ("ef4" in n or "ef 4" in n or "ef-4" in n or "ef_4" in n):
            return name
    for name in wb.sheetnames:
        n = _norm_text(name)
        if "ef-4"
